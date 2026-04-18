"""后台持续温湿度监控 - 追加到桌面 Excel"""
import serial
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import os
import re
import time
import sys

# 桌面路径
desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
xlsx_path = os.path.join(desktop, '温湿度记录.xlsx')

# 初始化或加载 Excel
if os.path.exists(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '温湿度记录'
    ws.append(['时间', '温度(C)', '湿度(%)'])
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    wb.save(xlsx_path)

print(f'Excel: {xlsx_path}')
print(f'当前行数: {ws.max_row}')
print('开始持续监控... (Ctrl+C 停止)\n')

# 状态文件
status_file = os.path.join(os.path.dirname(__file__), 'monitor_status.txt')

def write_status(msg):
    with open(status_file, 'w', encoding='utf-8') as f:
        f.write(f'{datetime.now()}: {msg}\n')

# 串口连接
ser = None
buffer = b''
count = 0
save_interval = 3  # 每3条保存一次

try:
    while True:
        # 尝试连接串口
        if ser is None or not ser.is_open:
            try:
                ser = serial.Serial('COM6', 9600, timeout=3)
                time.sleep(0.5)
                print(f'[{datetime.now().strftime("%H:%M:%S")}] 串口已连接')
                write_status('running')
            except Exception as e:
                print(f'[{datetime.now().strftime("%H:%M:%S")}] 串口连接失败: {e}')
                write_status(f'error: {e}')
                time.sleep(5)
                continue
        
        # 读取数据
        try:
            chunk = ser.read(100)
            if chunk:
                buffer += chunk
                
                # 解析数据
                while b'\r\n' in buffer:
                    line, buffer = buffer.split(b'\r\n', 1)
                    line = line.decode('ascii', errors='ignore')
                    
                    match = re.match(r'T:(\d+)C\s+H:(\d+)%', line)
                    if match:
                        temp = int(match.group(1))
                        humi = int(match.group(2))
                        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        ws.append([now, temp, humi])
                        count += 1
                        print(f'[{count}] {now}  T:{temp:2d}C  H:{humi:2d}%')
                        write_status(f'count={count}')
                        
                        # 定期保存
                        if count % save_interval == 0:
                            wb.save(xlsx_path)
                            # print(f'    (已保存 {ws.max_row} 行)')
                            
        except serial.SerialException as e:
            print(f'[{datetime.now().strftime("%H:%M:%S")}] 串口错误: {e}')
            write_status(f'serial error: {e}')
            try:
                ser.close()
            except:
                pass
            ser = None
            time.sleep(3)
            
except KeyboardInterrupt:
    print('\n用户停止')
    write_status('stopped')

finally:
    if ser:
        try:
            ser.close()
        except:
            pass
    wb.save(xlsx_path)
    print(f'\n完成! 共 {count} 条记录，总计 {ws.max_row} 行')
    print(f'文件: {xlsx_path}')
