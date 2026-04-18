"""实时温湿度监控 - 写入桌面 Excel"""
import serial
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os
import re
import time

# 桌面路径
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
xlsx_path = os.path.join(desktop, "温湿度记录.xlsx")

# 创建或加载工作簿
if os.path.exists(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    row = ws.max_row + 1
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "温湿度记录"
    # 表头
    ws.append(["时间", "温度(°C)", "湿度(%)"])
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    row = 2

print(f"Excel: {xlsx_path}")
print(f"从第 {row} 行开始写入\n")

# 串口
ser = serial.Serial('COM6', 9600, timeout=2)
time.sleep(1)

buffer = b''
count = 0
max_records = 30  # 记录30条后停止

print("开始监控... (Ctrl+C 停止)")
print("-" * 40)

try:
    while count < max_records:
        chunk = ser.read(100)
        if not chunk:
            continue
        buffer += chunk
        
        # 解析 T:xxC H:xx%
        while b'\r\n' in buffer:
            line, buffer = buffer.split(b'\r\n', 1)
            line = line.decode('ascii', errors='ignore')
            
            match = re.match(r'T:(\d+)C\s+H:(\d+)%', line)
            if match:
                temp = int(match.group(1))
                humi = int(match.group(2))
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                ws.append([now, temp, humi])
                count += 1
                print(f"[{now}] 温度: {temp}°C  湿度: {humi}%  → 已写入第 {row + count - 1} 行")
                
                # 每5条保存一次
                if count % 5 == 0:
                    wb.save(xlsx_path)
                    print(f"    (已保存)")
    
except KeyboardInterrupt:
    print("\n用户中断")

finally:
    ser.close()
    wb.save(xlsx_path)
    print(f"\n完成! 共写入 {count} 条记录")
    print(f"文件: {xlsx_path}")
